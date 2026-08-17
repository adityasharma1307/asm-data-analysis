"""
Builds ASM_Assignments_Dashboard.xlsx -- a single Excel dashboard summarizing
all three statistical analyses in this repo:
  Asm1 -- one-sample t-test on stroke-patient BMI
  Asm2 Part A -- multiple linear regression (advertising spend -> sales)
  Asm2 Part B -- one-way ANOVA + LDA classification (iris species)

The numbers below are the results printed by running:
  Asm1/Aditya_ASM_Assignment_1.py
  Asm2/advertising_regression.py
  Asm2/iris_anova_classification.py
Re-run those three scripts and update the numbers here (or replace with a
programmatic capture) if the underlying CSVs ever change.
"""
import os

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_PATH = os.path.join(BASE_DIR, "ASM_Assignments_Dashboard.xlsx")

FONT = "Calibri"
NAVY = "1F2937"
BLUE = "1565C0"
GREEN = "1B5E20"
AMBER = "B45309"
PURPLE = "6D28D9"

THIN = Border(*[Side(style="thin", color="D1D5DB")] * 4)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center")


def header_row(ws, row, ncols, fill=NAVY, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = CTR
        cell.border = THIN


def write_table(ws, headers, rows, start_row, start_col=1, fill=NAVY, number_cols=None):
    number_cols = number_cols or {}
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + j, value=h)
    header_row(ws, start_row, len(headers), fill=fill, start_col=start_col)
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=val)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = LFT if j == 0 else CTR
            cell.border = THIN
            if headers[j] in number_cols:
                cell.number_format = number_cols[headers[j]]
    return start_row + len(rows)


def autosize(ws, max_col, max_row, min_w=10, max_w=42):
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        best = min_w
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                best = max(best, len(str(v)) + 2)
        ws.column_dimensions[letter].width = min(best, max_w)


def kpi_card(ws, row, col, label, value, color):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
    lbl.fill = PatternFill("solid", fgColor=color)
    lbl.alignment = CTR
    val = ws.cell(row=row + 1, column=col, value=value)
    val.font = Font(name=FONT, size=16, bold=True, color=color)
    val.alignment = CTR
    for rr in (row, row + 1, row + 2):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).border = THIN


def main():
    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------------------------------------------ #
    # Dashboard sheet
    # ------------------------------------------------------------------ #
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws["B2"] = "MATH F432 -- Applied Statistical Methods"
    ws["B2"].font = Font(name=FONT, bold=True, size=18, color=NAVY)
    ws["B3"] = "Assignment 1 (BMI t-test) + Assignment 2 (Regression, ANOVA, Classification)"
    ws["B3"].font = Font(name=FONT, size=11, italic=True, color="6B7280")

    kpis = [
        ("Asm1: BMI t-test p-value", "< 0.001", BLUE),
        ("Asm2A: Regression R²", "0.903", GREEN),
        ("Asm2B: ANOVA effect size (η²)", "0.941", AMBER),
        ("Asm2B: LDA test accuracy", "97.8%", PURPLE),
    ]
    for i, (label, value, color) in enumerate(kpis):
        kpi_card(ws, 5, 2 + i * 3, label, value, color)

    # --- BMI t-test summary table + chart ---
    row0 = 10
    ws.cell(row=row0, column=2, value="Asm1 -- BMI One-Sample t-Test").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    last1 = write_table(
        ws, ["Metric", "Value"],
        [
            ["Sample mean BMI", 30.47], ["Hypothesized mean (H0)", 28.50],
            ["t-statistic", 4.503], ["Critical value (±)", 1.971],
            ["95% CI lower", 29.61], ["95% CI upper", 31.33],
            ["Cohen's d (effect size)", 0.311], ["Wilcoxon p-value", 0.00036],
        ],
        start_row=row0 + 1, start_col=2, fill=BLUE,
    )
    autosize(ws, 3, last1)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "BMI: Sample Mean vs. Hypothesized Mean (with 95% CI)"
    chart1.y_axis.title = "BMI (kg/m²)"
    means_row = row0 + 1
    data1 = Reference(ws, min_col=3, max_col=3, min_row=means_row, max_row=means_row + 2)
    cats1 = Reference(ws, min_col=2, min_row=means_row + 1, max_row=means_row + 2)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    chart1.width, chart1.height = 13, 8
    ws.add_chart(chart1, "B21")

    # --- Regression coefficients table + chart ---
    col0 = 7
    ws.cell(row=row0, column=col0, value="Asm2A -- Regression Coefficients").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    last2 = write_table(
        ws, ["Predictor", "Coefficient", "p-value"],
        [
            ["Intercept", 4.6251, 0.000],
            ["TV", 0.0544, 0.000],
            ["Radio", 0.1070, 0.000],
            ["Newspaper", 0.0003, 0.954],
        ],
        start_row=row0 + 1, start_col=col0, fill=GREEN,
    )
    autosize(ws, col0 + 2, last2)

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Regression Coefficients (Sales ~ TV + Radio + Newspaper)"
    chart2.y_axis.title = "Coefficient"
    data2 = Reference(ws, min_col=col0 + 1, max_col=col0 + 1, min_row=row0 + 2, max_row=row0 + 5)
    cats2 = Reference(ws, min_col=col0, min_row=row0 + 2, max_row=row0 + 5)
    chart2.add_data(data2, titles_from_data=False)
    chart2.set_categories(cats2)
    chart2.width, chart2.height = 13, 8
    ws.add_chart(chart2, "H21")

    # --- ANOVA group means table + chart ---
    row1 = 38
    ws.cell(row=row1, column=2, value="Asm2B -- Petal Length by Species (ANOVA)").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    last3 = write_table(
        ws, ["Species", "Mean Petal Length", "Std. Dev."],
        [
            ["setosa", 1.464, 0.174],
            ["versicolor", 4.260, 0.470],
            ["virginica", 5.552, 0.552],
        ],
        start_row=row1 + 1, start_col=2, fill=AMBER,
    )
    autosize(ws, 4, last3)

    chart3 = BarChart()
    chart3.type = "col"
    chart3.title = "Mean Petal Length by Species"
    chart3.y_axis.title = "Petal Length (cm)"
    data3 = Reference(ws, min_col=3, max_col=3, min_row=row1 + 1, max_row=row1 + 4)
    cats3 = Reference(ws, min_col=2, min_row=row1 + 2, max_row=row1 + 4)
    chart3.add_data(data3, titles_from_data=True)
    chart3.set_categories(cats3)
    chart3.width, chart3.height = 13, 8
    ws.add_chart(chart3, "B49")

    # --- LDA classification report table + chart ---
    ws.cell(row=row1, column=7, value="Asm2B -- LDA Classification Report").font = Font(name=FONT, bold=True, size=12, color=NAVY)
    last4 = write_table(
        ws, ["Species", "Precision", "Recall", "F1"],
        [
            ["setosa", 1.00, 1.00, 1.00],
            ["versicolor", 0.94, 1.00, 0.97],
            ["virginica", 1.00, 0.93, 0.97],
        ],
        start_row=row1 + 1, start_col=7, fill=PURPLE,
    )
    autosize(ws, 10, last4)

    chart4 = BarChart()
    chart4.type = "col"
    chart4.title = "LDA per-species F1-score (test accuracy = 97.8%)"
    chart4.y_axis.title = "F1-score"
    data4 = Reference(ws, min_col=10, max_col=10, min_row=row1 + 1, max_row=row1 + 4)
    cats4 = Reference(ws, min_col=7, min_row=row1 + 2, max_row=row1 + 4)
    chart4.add_data(data4, titles_from_data=True)
    chart4.set_categories(cats4)
    chart4.width, chart4.height = 13, 8
    ws.add_chart(chart4, "H49")

    ws.column_dimensions["A"].width = 2

    # ------------------------------------------------------------------ #
    # Detail sheets
    # ------------------------------------------------------------------ #
    s1 = wb.create_sheet("Asm1 BMI Test")
    write_table(
        s1, ["Metric", "Symbol", "Value"],
        [
            ["Sample size", "n", 209], ["Sample mean BMI", "x̄", 30.47],
            ["Sample std. dev.", "s", 6.33], ["Hypothesized mean", "μ0", 28.50],
            ["Standard error", "SE", 0.44], ["Degrees of freedom", "df", 208],
            ["t-statistic", "t", 4.503], ["Critical value (two-tailed, α=0.05)", "t*", 1.971],
            ["Margin of error", "ME", 0.863], ["95% CI lower", "", 29.61],
            ["95% CI upper", "", 31.33], ["Decision", "", "Reject H0"],
            ["Shapiro-Wilk statistic", "W", 0.9564], ["Shapiro-Wilk p-value", "p", "5.24e-06"],
            ["Cohen's d", "d", 0.311], ["Wilcoxon signed-rank p-value", "p", "3.60e-04"],
            ["Bootstrap 95% CI lower", "", 29.63], ["Bootstrap 95% CI upper", "", 31.33],
        ], start_row=1, fill=BLUE,
    )
    autosize(s1, 3, 19)

    s2 = wb.create_sheet("Asm2A Regression")
    r = write_table(
        s2, ["Predictor", "Coefficient", "Std. Error", "t", "p-value", "95% CI"],
        [
            ["Intercept", 4.6251, 0.308, 15.041, "< 0.001", "(4.019, 5.232)"],
            ["TV", 0.0544, 0.001, 39.592, "< 0.001", "(0.052, 0.057)"],
            ["Radio", 0.1070, 0.008, 12.604, "< 0.001", "(0.090, 0.124)"],
            ["Newspaper", 0.0003, 0.006, 0.058, 0.954, "(-0.011, 0.012)"],
        ], start_row=1, fill=GREEN,
    )
    r += 2
    s2.cell(row=r, column=1, value="Model fit / diagnostics").font = Font(name=FONT, bold=True, size=11)
    r += 1
    write_table(
        s2, ["Metric", "Value"],
        [
            ["R-squared", 0.903], ["Adjusted R-squared", 0.901],
            ["F-statistic (3, 196)", 605.4], ["F-test p-value", "8.13e-99"],
            ["VIF: TV", 2.487], ["VIF: Radio", 3.285], ["VIF: Newspaper", 3.055],
            ["Shapiro-Wilk on residuals (p)", 0.0016],
            ["Breusch-Pagan (p)", 0.2638],
        ], start_row=r, fill=GREEN,
    )
    autosize(s2, 6, r + 10)

    s3 = wb.create_sheet("Asm2B ANOVA + LDA")
    r = write_table(
        s3, ["Species", "n", "Mean Petal Length", "Std. Dev."],
        [
            ["setosa", 50, 1.464, 0.174],
            ["versicolor", 50, 4.260, 0.470],
            ["virginica", 50, 5.552, 0.552],
        ], start_row=1, fill=AMBER,
    )
    r += 2
    s3.cell(row=r, column=1, value="ANOVA & robustness checks").font = Font(name=FONT, bold=True, size=11)
    r += 1
    r = write_table(
        s3, ["Test", "Statistic", "p-value", "Conclusion"],
        [
            ["Classical one-way ANOVA", "F(2,147)=1179.03", "3.05e-91", "Reject H0"],
            ["Welch's ANOVA (unequal var.)", "F(2,78.05)=1826.58", "2.85e-66", "Reject H0"],
            ["Kruskal-Wallis (non-parametric)", "H=130.41", "4.80e-29", "Reject H0"],
            ["Levene's test (equal variances)", "stat=19.72", "2.59e-08", "Variances differ"],
        ], start_row=r, fill=AMBER,
    )
    r += 2
    s3.cell(row=r, column=1, value="Eta-squared (effect size): 0.941").font = Font(name=FONT, size=10, italic=True)
    r += 2
    s3.cell(row=r, column=1, value="LDA classification report (test accuracy = 97.8%)").font = Font(name=FONT, bold=True, size=11)
    r += 1
    write_table(
        s3, ["Species", "Precision", "Recall", "F1-score", "Support"],
        [
            ["setosa", 1.00, 1.00, 1.00, 15],
            ["versicolor", 0.94, 1.00, 0.97, 15],
            ["virginica", 1.00, 0.93, 0.97, 15],
        ], start_row=r, fill=PURPLE,
    )
    autosize(s3, 5, r + 6)

    wb.save(OUT_PATH)
    print(f"Dashboard saved: {OUT_PATH}")


if __name__ == "__main__":
    main()
